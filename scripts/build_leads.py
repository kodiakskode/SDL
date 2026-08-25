import json, re, unicodedata
import openpyxl

SRC = "/root/.claude/uploads/396cf451-05e8-5691-9770-c8e7641c822f/"

def org_clean(v):
    """Practice names are free text; a lone punctuation mark is not a name."""
    s = title(v)
    return "" if not re.search(r"[A-Za-z0-9]", s) else s

def clean(v):
    if v is None: return ""
    s = str(v).strip()
    s = unicodedata.normalize("NFKC", s)
    s = re.sub(r"\s+", " ", s)
    if s.lower() in ("none", "n/a", "na", "-", "—", "nan", ""): return ""
    return s

def phone(v):
    s = clean(v)
    if not s: return ""
    s = s.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
    if s.startswith("+61"): s = "0" + s[3:]
    if not re.fullmatch(r"[0-9]{6,12}", s): return clean(v)
    if s.startswith("04") and len(s) == 10:
        return f"{s[:4]} {s[4:7]} {s[7:]}"
    if s.startswith("1300") or s.startswith("1800"):
        return f"{s[:4]} {s[4:7]} {s[7:]}" if len(s) == 10 else s
    if s.startswith("02") and len(s) == 10:
        return f"{s[:2]} {s[2:6]} {s[6:]}"
    return s

def email(v):
    s = clean(v).lower()
    return s if re.fullmatch(r"[^@\s]+@[^@\s]+\.[a-z]{2,}", s) else ""

def website(v):
    s = clean(v)
    if not s: return ""
    s = s.rstrip("/")
    if not s.startswith("http"): s = "https://" + s.lstrip("/")
    return s

PC_REGIONS = [
    (2000, 2010, "Sydney CBD & Surrounds"),
    (2011, 2036, "Eastern Suburbs"),
    (2037, 2050, "Inner West"),
    (2060, 2091, "Lower North Shore"),
    (2092, 2108, "Northern Beaches"),
    (2110, 2126, "Upper North Shore & Hills"),
    (2127, 2145, "Parramatta & Inner West"),
    (2146, 2179, "Western Sydney"),
    (2190, 2234, "St George & Sutherland"),
    (2250, 2263, "Central Coast"),
    (2264, 2340, "Hunter & Newcastle"),
    (2350, 2490, "Northern NSW & New England"),
    (2500, 2540, "Illawarra & South Coast"),
    (2541, 2620, "Southern Tablelands"),
    (2621, 2739, "Riverina & Murray"),
    (2740, 2786, "Blue Mountains, Penrith & Hawkesbury"),
    (2787, 2899, "Central West & Far West"),
]

def region_from_pc(pc):
    if not pc or not pc.isdigit(): return ""
    n = int(pc)
    for lo, hi, name in PC_REGIONS:
        if lo <= n <= hi: return name
    return "Regional NSW"

def split_address(addr):
    """'15 Lesley Close, Elanora Heights NSW 2101' -> (suburb, postcode)."""
    a = clean(addr)
    if not a: return "", ""
    pc = ""
    m = re.search(r"\b(2\d{3})\b\s*$", a)
    if m:
        pc = m.group(1)
        a = a[:m.start()].strip(" ,")
    a = re.sub(r"[,\s]+NSW$", "", a, flags=re.I).strip(" ,")
    part = a.split(",")[-1].strip()
    part = re.sub(r"^(shop|suite|unit|level|po box)\b.*", "", part, flags=re.I).strip()
    # a trailing street-number fragment is not a suburb
    if re.fullmatch(r"[\d/\-]+", part): part = ""
    return title(part), pc

def title(v):
    s = clean(v)
    if not s: return ""
    if s.isupper() or s.islower():
        s = " ".join(w.capitalize() for w in s.split(" "))
        for small in ("Of", "The", "And", "De", "Van"):
            s = re.sub(rf"(?<! )\b{small}\b", small.lower(), s)
    return s

# ── Architects ────────────────────────────────────────────────────
wb = openpyxl.load_workbook(SRC + "6311744f-NSW_Architects_750_Contacts.xlsx",
                            read_only=True, data_only=True)
ws = wb["NSW Architects"]
rows = list(ws.iter_rows(values_only=True))
wb.close()
hdr_i = next(i for i, r in enumerate(rows) if r and r[0] == "#")
arch, seen = [], set()
for r in rows[hdr_i + 1:]:
    if not r or r[0] is None: continue
    name = title(r[1])
    if not name: continue
    key = (name.lower(), clean(r[2]))
    if key in seen: continue
    seen.add(key)
    arch.append({
        "name": name,
        "org": org_clean(r[6]),
        "suburb": title(r[4]),
        "postcode": clean(r[5]),
        "region": region_from_pc(clean(r[5])),
        "email": email(r[8]),
        "phone": phone(r[9]),
        "website": website(r[7]),
        "ref": clean(r[2]),
        "type": "Architect",
        "status": clean(r[3]),
        "statusNote": "NSW ARB registration " + clean(r[2]),
        "source": website(r[11]),
    })

# ── Landscape designers ───────────────────────────────────────────
wb = openpyxl.load_workbook(SRC + "3c2dec67-NSW_Landscape_Designer_Contacts_750.xlsx",
                            read_only=True, data_only=True)
ws = wb["750 NSW Contacts"]
rows = list(ws.iter_rows(values_only=True))
wb.close()
hdr_i = next(i for i, r in enumerate(rows) if r and r[0] == "#")
land, seen = [], set()
for r in rows[hdr_i + 1:]:
    if not r or r[0] is None: continue
    org = org_clean(r[2])
    name = title(r[1])
    if not (org or name): continue
    key = (org.lower(), name.lower())
    if key in seen: continue
    seen.add(key)
    raw_status = clean(r[9])
    low = raw_status.lower()
    accredited = bool(re.search(r"\b(LDI|AILDM)\b", raw_status)) and "not independently verified" not in low and "unverified" not in low
    addr = clean(r[6])
    sub, pc = split_address(addr)
    reg = clean(r[7]) or region_from_pc(pc)
    land.append({
        "name": name,
        "org": org,
        "suburb": sub,
        "postcode": pc,
        "region": reg,
        "email": email(r[4]),
        "phone": phone(r[3]),
        "website": website(r[5]),
        "ref": "",
        "type": clean(r[8]),
        "status": "LDI / AILDM evidenced" if accredited else "Public listing — unverified",
        "statusNote": raw_status,
        "source": website(r[10]),
    })

def stats(rows):
    return {
        "total": len(rows),
        "withEmail": sum(1 for r in rows if r["email"]),
        "withPhone": sum(1 for r in rows if r["phone"]),
        "withWeb": sum(1 for r in rows if r["website"]),
    }

out = {
    "architects": {"label": "Architects",
                   "note": "NSW Architects Registration Board public profiles, compiled 24 Aug 2026.",
                   "rows": arch, "stats": stats(arch)},
    "designers": {"label": "Landscape Designers",
                  "note": "NSW landscape / garden design businesses from public listings; LDI / AILDM membership marked only where separately evidenced.",
                  "rows": land, "stats": stats(land)},
}
with open("/home/user/SDL/data/leads.json", "w") as f:
    json.dump(out, f, separators=(",", ":"), ensure_ascii=False)

for k, v in out.items():
    print(k, v["stats"])
print("sample arch:", json.dumps(arch[0], ensure_ascii=False))
print("sample land:", json.dumps(land[0], ensure_ascii=False))
